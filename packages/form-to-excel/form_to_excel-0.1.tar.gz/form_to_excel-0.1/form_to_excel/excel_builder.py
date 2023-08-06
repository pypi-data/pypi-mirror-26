from openpyxl import Workbook

class ExcelBuilder:
  def __init__(self, text):
    self.filename = text["filename"]
    self.questions = self.create_questions(text["question_labels"])
    self.wb = Workbook()
    self.ws = self.wb.worksheets[0]

  def generate_excel_file(self):
    for row, vals in enumerate(self.questions):
      column_cell_A = 'A'
      column_cell_B = 'B'
      self.ws[column_cell_A + str(row + 1)] = str(vals[0])
      self.ws[column_cell_B + str(row + 1)] = str(vals[1])

    self.wb.save("{}.xlsx".format(self.filename))

  def create_questions(self, item):
    questions = []
    for question in item:
      lowered = question.lower()
      if 'date' in lowered and 'dates' not in lowered and 'time' not in lowered:
        question_type = 'dt'
      elif 'time' in lowered and 'times' not in lowered and 'date' not in lowered:
        question_type = 'tm'
      elif 'time' in lowered and 'date' in lowered:
        question_type = 'dtm'
      elif 'latitude' in lowered or 'longitude' in lowered:
        question_type = 'loc'
      else:
        question_type = 'txt'
      questions.append((question_type, question))
    return questions
